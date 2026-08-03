from typing import Optional, Dict, Any, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, func, text
from datetime import datetime, date, timedelta
import uuid
import json
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from app.models.booking import Booking, BookingStatus
from app.models.payment import Payment, PaymentStatus
from app.models.schedule import Schedule
from app.models.vehicle import Vehicle
from app.models.driver import Driver
from app.models.customer import Customer
from app.models.schedule import ScheduleStatus


class ReportService:
    @staticmethod
    async def generate_booking_report(
        db: AsyncSession,
        start_date: date,
        end_date: date,
        format: str = "json"
    ) -> Dict[str, Any]:
        """
        Generate booking report for a date range
        """
        # Get bookings
        query = select(Booking).where(
            and_(
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date,
                Booking.is_deleted.is_(None)
            )
        )
        result = await db.execute(query)
        bookings = result.scalars().all()

        # Calculate statistics
        total_bookings = len(bookings)
        confirmed = len([b for b in bookings if b.status == BookingStatus.CONFIRMED])
        cancelled = len([b for b in bookings if b.status == BookingStatus.CANCELLED])
        completed = len([b for b in bookings if b.status == BookingStatus.COMPLETED])

        total_revenue = sum(float(b.total_amount) for b in bookings if b.status != BookingStatus.CANCELLED)

        # Group by route
        route_stats = {}
        for booking in bookings:
            if booking.schedule and booking.schedule.route:
                route_name = f"{booking.schedule.route.origin} → {booking.schedule.route.destination}"
                if route_name not in route_stats:
                    route_stats[route_name] = {"count": 0, "revenue": 0}
                route_stats[route_name]["count"] += 1
                if booking.status != BookingStatus.CANCELLED:
                    route_stats[route_name]["revenue"] += float(booking.total_amount)

        report = {
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "summary": {
                "total_bookings": total_bookings,
                "confirmed": confirmed,
                "cancelled": cancelled,
                "completed": completed,
                "total_revenue": total_revenue,
                "average_booking_value": total_revenue / total_bookings if total_bookings > 0 else 0
            },
            "by_route": [
                {
                    "route": route,
                    "count": stats["count"],
                    "revenue": stats["revenue"]
                }
                for route, stats in route_stats.items()
            ],
            "bookings": [
                {
                    "reference": b.reference,
                    "customer": b.customer.full_name if b.customer else "",
                    "route": f"{b.schedule.route.origin} → {b.schedule.route.destination}" if b.schedule and b.schedule.route else "",
                    "amount": float(b.total_amount),
                    "status": b.status.value,
                    "date": b.booking_date.isoformat()
                }
                for b in bookings[:100]  # Limit for performance
            ]
        }

        return report

    @staticmethod
    async def generate_financial_report(
        db: AsyncSession,
        start_date: date,
        end_date: date,
        format: str = "json"
    ) -> Dict[str, Any]:
        """Generate financial report"""
        # Get payments
        query = select(Payment).where(
            and_(
                Payment.payment_date >= start_date,
                Payment.payment_date <= end_date,
                Payment.status == PaymentStatus.PAID,
                Payment.is_deleted.is_(None)
            )
        )
        result = await db.execute(query)
        payments = result.scalars().all()

        total_payments = sum(float(p.amount_paid) for p in payments)

        # By method
        method_stats = {}
        for payment in payments:
            method = payment.method
            if method not in method_stats:
                method_stats[method] = {"count": 0, "total": 0}
            method_stats[method]["count"] += 1
            method_stats[method]["total"] += float(payment.amount_paid)

        # Outstanding payments
        outstanding = await db.execute(
            select(func.sum(Booking.outstanding_balance))
            .where(
                and_(
                    Booking.outstanding_balance > 0,
                    Booking.status != BookingStatus.CANCELLED,
                    Booking.is_deleted.is_(None)
                )
            )
        )
        outstanding = outstanding.scalar() or 0

        report = {
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "summary": {
                "total_revenue": total_payments,
                "outstanding_payments": float(outstanding),
                "total_transactions": len(payments)
            },
            "by_method": [
                {
                    "method": method,
                    "count": stats["count"],
                    "total": stats["total"]
                }
                for method, stats in method_stats.items()
            ]
        }

        return report

    @staticmethod
    async def generate_fleet_report(
        db: AsyncSession
    ) -> Dict[str, Any]:
        """Generate fleet utilization report"""
        # Get all vehicles
        query = select(Vehicle).where(Vehicle.is_deleted.is_(None))
        result = await db.execute(query)
        vehicles = result.scalars().all()

        fleet_stats = []
        for vehicle in vehicles:
            # Get schedules for this vehicle
            schedules = await db.execute(
                select(Schedule).where(
                    and_(
                        Schedule.vehicle_id == vehicle.id,
                        Schedule.status != ScheduleStatus.CANCELLED,
                        Schedule.is_deleted.is_(None)
                    )
                )
            )
            schedules = schedules.scalars().all()

            total_trips = len(schedules)
            completed_trips = len([s for s in schedules if s.status == ScheduleStatus.COMPLETED])
            total_passengers = sum(s.booked_seats for s in schedules)
            utilization = (total_passengers / (vehicle.capacity * total_trips)) * 100 if total_trips > 0 else 0

            fleet_stats.append({
                "vehicle": vehicle.name or vehicle.registration,
                "type": vehicle.vehicle_type.value if vehicle.vehicle_type else "",
                "capacity": vehicle.capacity,
                "total_trips": total_trips,
                "completed_trips": completed_trips,
                "total_passengers": total_passengers,
                "utilization_rate": round(utilization, 2)
            })

        return {
            "total_vehicles": len(vehicles),
            "fleet_stats": fleet_stats
        }

    @staticmethod
    async def generate_customer_report(
        db: AsyncSession,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate customer report"""
        # Get customers with booking count
        customers = await db.execute(
            select(
                Customer,
                func.count(Booking.id).label("booking_count"),
                func.sum(Booking.total_amount).label("total_spent")
            )
            .outerjoin(Booking, and_(
                Booking.customer_id == Customer.id,
                Booking.booking_date.between(start_date, end_date),
                Booking.status != BookingStatus.CANCELLED,
                Booking.is_deleted.is_(None)
            ))
            .where(
                and_(
                    Customer.is_active == True,
                    Customer.is_deleted.is_(None)
                )
            )
            .group_by(Customer.id)
            .order_by(text("total_spent DESC"))
            .limit(100)
        )
        customers = customers.all()

        return {
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "total_customers": len(customers),
            "top_customers": [
                {
                    "name": c.Customer.full_name,
                    "phone": c.Customer.phone,
                    "email": c.Customer.email,
                    "bookings": c.booking_count or 0,
                    "total_spent": float(c.total_spent) if c.total_spent else 0
                }
                for c in customers
            ]
        }

    @staticmethod
    async def export_to_excel(
        data: Dict[str, Any],
        report_type: str
    ) -> BytesIO:
        """Export report to Excel format"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.title()

        # Add headers
        headers = []
        if report_type == "booking":
            headers = ["Reference", "Customer", "Route", "Amount", "Status", "Date"]
        elif report_type == "financial":
            headers = ["Method", "Count", "Total"]
        elif report_type == "fleet":
            headers = ["Vehicle", "Type", "Capacity", "Trips", "Completed", "Passengers", "Utilization"]
        elif report_type == "customer":
            headers = ["Name", "Phone", "Email", "Bookings", "Total Spent"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Add data
        row = 2
        if report_type == "booking" and "bookings" in data:
            for booking in data["bookings"]:
                ws.cell(row=row, column=1, value=booking.get("reference", ""))
                ws.cell(row=row, column=2, value=booking.get("customer", ""))
                ws.cell(row=row, column=3, value=booking.get("route", ""))
                ws.cell(row=row, column=4, value=booking.get("amount", 0))
                ws.cell(row=row, column=5, value=booking.get("status", ""))
                ws.cell(row=row, column=6, value=booking.get("date", ""))
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def export_to_csv(
        data: Dict[str, Any],
        report_type: str
    ) -> str:
        """Export report to CSV format"""
        output = StringIO()
        writer = csv.writer(output)

        # Write headers
        if report_type == "booking":
            writer.writerow(["Reference", "Customer", "Route", "Amount", "Status", "Date"])
            for booking in data.get("bookings", []):
                writer.writerow([
                    booking.get("reference", ""),
                    booking.get("customer", ""),
                    booking.get("route", ""),
                    booking.get("amount", 0),
                    booking.get("status", ""),
                    booking.get("date", "")
                ])
        elif report_type == "customer":
            writer.writerow(["Name", "Phone", "Email", "Bookings", "Total Spent"])
            for customer in data.get("top_customers", []):
                writer.writerow([
                    customer.get("name", ""),
                    customer.get("phone", ""),
                    customer.get("email", ""),
                    customer.get("bookings", 0),
                    customer.get("total_spent", 0)
                ])

        return output.getvalue()